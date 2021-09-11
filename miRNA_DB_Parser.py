import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time

in_values = pd.read_excel('derg_de_mirna.xlsx', sheet_name='Ls - Lsk')['name']
url_miRDB = 'http://mirdb.org/cgi-bin/search.cgi'

with pd.ExcelWriter('miRNA_output_table.xlsx', engine='openpyxl', mode='a') as writer:
    book = load_workbook('miRNA_output_table.xlsx')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sheet = book.active
    print('Writing to the outfile...')
    num = 0
    for value in in_values: #for each miRNA from input file
        try:    
            num +=1
            
            payload = {'species': 'Rat',        #
                       'searchBox': f"{value}", # data for the POST event
                       'submitButton': 'Go',    #
                       'searchType': 'miRNA'}   #
            resp_miRDB = requests.post(url_miRDB, data=payload)
            soup = BeautifulSoup(resp_miRDB.text, 'html.parser')
            table = soup.find_all('table')[1]
            element = table.find_all('tr') 
            
            pd_df = pd.DataFrame({'miRNA':[],
                                  'Rank':[],
                                  'Gene_Symbol':[],
                                  'Gene_Desciption':[]})
            for i in range(len(element)): 
                if i == 0: #skipping the head of each table
                    continue
                td_row = element[i].find_all('td')
                list_ = []
                for row in td_row:
                    list_.append(row.string.strip())
                if int(list_[2]) >= 90:
                    pd_df1 = pd.DataFrame({'miRNA': [list_[3]],
                                           'Target score': [int(list_[2])],
                                           'Gene_Symbol': [list_[4]],
                                           'Gene_Desciption': [list_[5]]})
                    pd_df = pd_df.append(pd_df1, ignore_index=True)           
            pd_df.to_excel(writer, sheet_name='miRDB', header=False, index=False, startrow=sheet.max_row+1)
            book.save('miRNA_output_table.xlsx')
        except AttributeError:
            continue
        print(f"{num}/{len(in_values)}", end='\r')    
        time.sleep(0.5)
print('Done!')
